from openpyxl import load_workbook


# 1.创建一个类
class ExcelUtil:

    # 1.1.初始化excel文件
    def __init__(self,excelf):
        # 获取excel路径
        self.excelf=excelf
    # 1.2打开excel文件
        self.workbook_obj=load_workbook(self.excelf)
    # 1.3读取excel文件对应的工作表sheet
        self.sheet_obj=self.workbook_obj["Sheet1"]

    #    2.定义一个方法方便调用
    def data(self):
        # 2.1读取工作表sheet相关数据，以列表的形式，每行显示
        self.result_all=list(self.sheet_obj.iter_rows(values_only=True))
        # 2.2读取excel表头信息
        excel_title=self.result_all[0]
        # print("表头信息为",excel_title)
        # 2.3读取excel表数据
        excel_data=self.result_all[1:]
        # print("表数据为",excel_data)
        # 3.获取目标列表字典
        target_list=[]
        for i in excel_data:
            # 3.1将表头和表数据变成一个字典
            dict_data=dict(zip(excel_title,i))
            # print(dict_data)
            # 3.2将字典变成一个列表字典
            target_list.append(dict_data)
        # print(target_list)
        return target_list
    # 4.关闭文件
    def CloseFile(self):
        self.excelf.close()

if __name__ == '__main__':
    E=ExcelUtil(r"G:\pythonProject2\10.20pytest\data\excel\excel_data.xlsx")
    print(E.data())
